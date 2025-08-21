import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from config import ISSUE_DETAILS, HEADER_COLOR
import datetime
from urllib.parse import urlparse
import re

def generate_xlsx_report(issues, base_url, crawled_count):
    """Analyzes the data and generates a styled report in XLSX format."""
    # Sanitize base_url for filename
    parsed_url = urlparse(base_url)
    domain_name = parsed_url.netloc.replace("www.", "").replace(".", "_") # Replace dots with underscores
    domain_name = re.sub(r'[^a-zA-Z0-9_ -]', '', domain_name) # Remove other invalid characters
    
    # Get current timestamp
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    
    report_filename = f"seo_audit_report_{domain_name}_{timestamp}.xlsx"
    with pd.ExcelWriter(report_filename, engine='openpyxl') as writer:
        # --- Styles ---
        header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # --- Summary Sheet ---
        summary_list = [("Metric", "Value"), ("Base URL", base_url), ("Pages Crawled", crawled_count), ("", "")]
        for issue_key, details in ISSUE_DETAILS.items():
            if issues.get(issue_key):
                summary_list.append((details["sheet_name"], len(issues[issue_key])))
        
        summary_df = pd.DataFrame(summary_list[1:], columns=summary_list[0])
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        ws = writer.sheets["Summary"]
        for cell in ws["A"] + ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # --- Issue Sheets ---
        for issue_key, details in ISSUE_DETAILS.items():
            data = issues.get(issue_key)
            if not data: continue

            sheet_name = details["sheet_name"]
            info_df = pd.DataFrame([{"Info": details["description"]}, {"Info": details["recommendation"]}])
            info_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=0)

            # Create DataFrame from issue data and add 'Status' column
            issue_df = pd.DataFrame(data)
            issue_df["Status"] = "Pending" # Initialize with default status
            
            if not issue_df.empty:
                issue_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)
                ws = writer.sheets[sheet_name]
                
                # Define status options for data validation
                status_options = ['Pending', 'In Progress', 'Completed']

                # Apply Data Validation for 'Status' column
                # Assuming 'Status' column is the last one added (index len(issue_df.columns) - 1)
                status_col_index = len(issue_df.columns)
                status_col_letter = chr(ord('A') + status_col_index - 1)
                
                # Start from row 5 (after header, description, recommendation, and data header)
                start_row_data = 5
                end_row_data = len(issue_df) + start_row_data - 1

                # Define the range for data validation using sqref
                data_validation_range = f'{status_col_letter}{start_row_data}:{status_col_letter}{end_row_data}'
                dv = DataValidation(type="list", formula1=f'"{",".join(status_options)}"', showDropDown=True, sqref=data_validation_range)

                # Add data validation to the worksheet for this sheet
                ws.add_data_validation(dv)

                # Style headers
                for cell in ws[1] + ws[2]: # Description and Recommendation
                    cell.fill = header_fill
                    cell.font = header_font
                for cell in ws[4]: # Data headers
                    cell.fill = header_fill
                    cell.font = header_font

        # --- Auto-adjust column widths for all sheets ---
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column_cells in worksheet.columns:
                max_length = 0
                for cell in column_cells:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    print(f"\nSEO analysis XLSX report saved to {report_filename}")
